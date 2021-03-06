import numpy as np
import matplotlib as mpl
import seaborn as sns
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def create_results_directory(results_directory, folders=None, excels=None):
    if os.path.exists(results_directory):
        expand = 1
        while True:
            expand += 1
            new_results_directory = results_directory + '_' + str(expand)
            if os.path.exists(new_results_directory):
                continue
            else:
                results_directory = new_results_directory
                break
    os.mkdir(results_directory)

    if folders:
        for item in folders:
            os.mkdir(results_directory + '/' + item)

    if excels:
        for item in excels:
            if item[-5:] != '.xlsx':
                item = item + '.xlsx'
            excel_name = results_directory + '/' + item
            wb = openpyxl.Workbook()
            wb.save(excel_name)
            wb.close()

    print('Creating new results directory: {}'.format(results_directory))
    return results_directory


def create_excel_file(excel_name):
    while os.path.isfile(excel_name):
        expand = 1
        while True:
            expand += 1
            new_file_name = excel_name.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                excel_name = new_file_name
                break
    print('Writing into' + excel_name + '\n')
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    return excel_name


def print_df_to_excel(df, ws, start_row=1, start_col=1, index=True, header=True):
    rows = list(dataframe_to_rows(df, index=index, header=header))
    rows.pop(1)
    for r_idx, row in enumerate(rows, start_row):
        skip_count = 0
        for c_idx, value in enumerate(row, start_col):
            if isinstance(value, str):
                if 'Unnamed' not in value:
                    ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
            else:
                ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
        else:
            skip_count += 1


def print_array_to_excel(array, first_cell, ws, axis=2):
    '''
    Print an np array to excel using openpyxl
    :param array: np array
    :param first_cell: first cell to start dumping values in
    :param ws: worksheet reference. From openpyxl, ws=wb[sheetname]
    :param axis: to determine if the array is a col vector (0), row vector (1), or 2d matrix (2)
    '''
    if isinstance(array, (list,)):
        array = np.array(array)
    shape = array.shape
    if axis == 0:
        # Treat array as col vector and print along the rows
        array.flatten()  # Flatten in case the input array is a nx1 ndarry which acts weird
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i]
    elif axis == 1:
        # Treat array as row vector and print along the columns
        array.flatten()  # Flatten in case the input array is a 1xn ndarry which acts weird
        for j in range(shape[0]):
            i = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[j]
    elif axis == 2:
        # If axis==2, means it is a 2d array
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]


def create_id_dict(var_name, h, expt, est, model=None, model_name=None, seed=None,  combined_name=False, **kwargs):
    if combined_name:
       return {'var_name': var_name, 'model_full_name':f'{var_name}_{combined_name}','model': combined_name,
               'model_name': combined_name, 'est':est, 'seed':seed,
               'results_dir':f'./results/{expt}/model_combination_{var_name}_{combined_name}'}

    if not est:
        results_dir = f'./results/{expt}/poos_{var_name}_{model_name}'
        model_full_name = f'{var_name}_{model_name}'
        model_full_name_wo_var = model_name
    else:
        results_dir = f'./results/{expt}/poos_{var_name}_{model_name}_{est}_s{seed}'
        model_full_name = f'{var_name}_{model_name}_{est}_s{seed}'
        model_full_name_wo_var = f'{model_name}_{est}_s{seed}'
    return {'var_name': var_name, 'h': h, 'est': est, 'model': model, 'model_name': model_name, 'expt': expt,
               'seed': seed, 'model_full_name_wo_var': model_full_name_wo_var,
               'results_dir': results_dir,
               'model_full_name': model_full_name,}


def create_id_store(var_name,expt_type, est,  model, model_name,  seed):
    return [create_id_dict(var_name=var_name,
                           h=[1, 3, 6, 12, 24],
                           est=e,
                           model=m,
                           model_name=mn,
                           expt=expt_type,
                           seed=s)
            for e, m, mn, s in zip(est, model, model_name, seed)]


def set_matplotlib_style():
    try:
        del mpl.font_manager.weight_dict['roman']
        mpl.font_manager._rebuild()
    except KeyError:
        pass
    sns.set(style='ticks')
    mpl.rc('font', family='Times New Roman')

